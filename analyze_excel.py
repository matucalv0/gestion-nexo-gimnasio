import openpyxl

file_path = r"src\main\resources\bd\Mati Lacerenze.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb.active

print("="*80)
print("ESTRUCTURA COMPLETA DEL EXCEL DE RUTINAS")
print("="*80)

print(f"\nTotal filas: {ws.max_row}, Total columnas: {ws.max_column}")

print("\nPrimeras 20 filas (completas):")
for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
    non_empty = [str(v) if v is not None else '' for v in row]
    print(f"  Fila {row_idx:2d}: {' | '.join(non_empty[:9])}")

# Analizar estructura de días
print("\n" + "="*80)
print("ESTRUCTURA DE DATOS:")
print("="*80)

current_day = None
dias = {}

for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True), start=1):
    first_cell = row[0]
    
    if first_cell and isinstance(first_cell, str) and 'Dia' in first_cell:
        current_day = first_cell
        dias[current_day] = []
        print(f"\n{current_day}")
    elif first_cell and isinstance(first_cell, str) and first_cell.strip() == 'Ejercicio':
        print(f"  Encabezados: {[v for v in row if v]}")
    elif current_day and first_cell and isinstance(first_cell, str) and first_cell not in ['Dia', 'Nombre', 'Fecha']:
        # Es un ejercicio
        dias[current_day].append(row)
        print(f"  → {row[:9]}")

print("\n" + "="*80)
print("RESUMEN:")
print("="*80)
for day, exercises in dias.items():
    print(f"{day}: {len(exercises)} ejercicios")
